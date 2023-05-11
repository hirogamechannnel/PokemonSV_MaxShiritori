import setup
import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
import datetime
import openpyxl
from openpyxl.styles import *
import pykakasi
import pulp
import networkx as nx
from ortoolpy import *

#F01 DataFrame生成用
def Mk_DataFrame(ListA,ListB):
    Df = pd.DataFrame(data=[ListA,ListB])
    Df = Df.T
    Df.columns = ['Zukan','PokeName']
    Df = Df.dropna(how='any')
    Df.index = ListA
    return Df

#1 前回ファイルの削除
def Rem_File():
    if os.path.exists(setup.ExcelFile) == True:
        os.remove(setup.ExcelFile)

#2 ポケモンwikiからポケモン名のリスト抽出
def Get_PokeName():

    print('#2-1_ポケモンwikiからデータ抽出')
    html = requests.get(setup.Poke_List_URL)
    html.encoding = html.apparent_encoding

    soup = BeautifulSoup(html.content,"html.parser")
    print(soup)

    chap1 = soup.find(class_ ="mw-parser-output")
    chap2 = chap1.find('table')

    for t in chap2.find_all('tr'):

        for s in t.find_all('a'):
            if s.text not in setup.Poke_Name_List and s.text not in setup.Poke_Except_List:
                setup.Poke_Name_List.append(s.text.replace('\n',''))


        for s in t.find_all('td'):
            if str.isdigit(s.text.replace('\n','')):
                if int(s.text) not in setup.Poke_Num_List:
                    setup.Poke_Num_List.append(int(s.text))
    
    Poke_Df = Mk_DataFrame(setup.Poke_Num_List,setup.Poke_Name_List)
    print('#2-2_ポケモン数:' + str(len(setup.Poke_Num_List)))

    return Poke_Df

#3 最後に「ん」の付くポケモンを除外
def Rem_Poke_LastN(Rep_Poke_Df):

    Name_List = []
    Num_List = []

    for t,row in Rep_Poke_Df.iterrows():
        LastTxt = row['PokeName'][len(row['PokeName']) -1]

        if LastTxt != 'ン':
            Num_List.append(row['Zukan'])
            Name_List.append(row['PokeName'])
        else:
            print('#3-1_除外ポケモン:' + row['PokeName'])

    Df = Mk_DataFrame(Num_List,Name_List)
    print('#3-2_しりとり対象ポケモン数:' + str(len(Num_List)))
    return Df

#4 最後の文字が特殊文字の場合にカタカナに変更
def Rep_SpecialtoChara(Rep_Poke_Df):

    Name_List = []
    Num_List = []

    for r,row in Rep_Poke_Df.iterrows():
        LastTxt = row['PokeName'][-1]

        if LastTxt in setup.Bef_Special_List:
            index = setup.Bef_Special_List.index(LastTxt)
            NewPokeNM = row['PokeName'][:-1] + setup.Aft_Special_List[index]
            print('#4-1_特殊文字変更ポケモン:' + NewPokeNM)
        else:
            NewPokeNM = row['PokeName']

        Num_List.append(row['Zukan'])
        Name_List.append(NewPokeNM)
    
    Df = Mk_DataFrame(Num_List,Name_List)
    return Df
        
#5 最後の文字がハイフンの場合の処理
def Rep_HyphentoChara(Rep_Poke_Df,Hyphen_JD):

    Name_List = []
    Num_List = []
    Roman = pykakasi.kakasi()

    for r,row in Rep_Poke_Df.iterrows():
        LastTxt = row['PokeName'][-1]
        Last2Txt = row['PokeName'][-2]
        
        if LastTxt == 'ー':
            if Hyphen_JD:
                NewPokeNM = row['PokeName'][:-1] + Last2Txt
                print('#5-T_ハイフンを直前文字に変更ポケモン:' + NewPokeNM)
            else:
                RomanResult = Roman.convert(row['PokeName'])
                for s in RomanResult:
                    RomaTxt = s['hepburn'][-1]
                    if RomaTxt == 'a':
                        NewTxt = 'ア'
                    elif RomaTxt == 'i':
                        NewTxt = 'イ'
                    elif RomaTxt == 'u':
                        NewTxt = 'ウ'
                    elif RomaTxt == 'e':
                        NewTxt = 'エ'
                    elif RomaTxt == 'o':
                        NewTxt = 'オ'
                NewPokeNM = row['PokeName'][:-1] + NewTxt
                print('#5-F_ハイフンを直前文字の母音に変更ポケモン:' + NewPokeNM)       

        else:
            NewPokeNM = row['PokeName']
        
        Num_List.append(row['Zukan'])
        Name_List.append(NewPokeNM)
        Df = Mk_DataFrame(Num_List,Name_List)
        
    return Df

#6 最後の文字が小文字の場合に大文字に変更
def Rep_SmalltoBig(Rep_Poke_Df):

    Name_List = []
    Num_List = []

    for r,row in Rep_Poke_Df.iterrows():
        LastTxt = row['PokeName'][-1]

        if LastTxt in setup.SmallChara_List:
            index = setup.SmallChara_List.index(LastTxt)
            NewPokeNM = row['PokeName'][:-1] + setup.BigChara_List[index]
            print('#6-1_小文字→大文字変更ポケモン:' + NewPokeNM)
        else:
            NewPokeNM = row['PokeName']
    
        Num_List.append(row['Zukan'])
        Name_List.append(NewPokeNM)
    
    Df = Mk_DataFrame(Num_List,Name_List)
    return Df

#7 濁点を考慮しない場合の濁点除去処理
def Rem_Dakuten(Rep_Poke_Df):
    
    Name_List = []
    Num_List = []

    for r,row in Rep_Poke_Df.iterrows():
        FirstTxt = row['PokeName'][0]
        LastTxt = row['PokeName'][-1]
        JD = False
        NewPokeNM = row['PokeName']

        if FirstTxt in setup.DakuKana:
            index = setup.DakuKana.find(FirstTxt)
            NewPokeNM = setup.Kana[index] + NewPokeNM[1:]
            JD = True
        
        if LastTxt in setup.DakuKana:
            index = setup.DakuKana.find(LastTxt)
            NewPokeNM = NewPokeNM[:-1] + setup.Kana[index]
            JD = True

        if JD:
            print('#7-T_濁点除去ポケモン:' + NewPokeNM)

        Num_List.append(row['Zukan'])
        Name_List.append(NewPokeNM)
    
    Df = Mk_DataFrame(Num_List,Name_List)

    return Df

#8 しりとり作成
def Mk_Shiritori(Rep_Poke_Df,Poke_Df):
    
    kws = []

    for r in Rep_Poke_Df['PokeName']:
        kws.append(r)

    g = nx.MultiDiGraph()
    g.add_nodes_from(['start','end'])

    for r in kws:
        g.add_edge(r[0],r[-1],word = r,var = addbinvar())

    for s in list(g.nodes)[2:]:
        g.add_edge('start',s,word = '' ,var = addbinvar())
        g.add_edge(s,'end',word = '' , var = addbinvar())

    df = pd.DataFrame([(fr,to,k,d['word'],d['var'])
        for (fr,to,k),d in g.edges.items()],
        columns = ['From','To','Key','Word','Var'])
    
    m = pulp.LpProblem(sense=LpMaximize)
    m += lpSum(df['Var'])#目的関数
    m += lpSum(df[df['From'] == 'start']['Var']) == 1
    m += lpSum(df[df['To'] == 'end']['Var']) == 1

    for s in list(g.nodes())[2:]:
        m += (lpSum([t[2] for t in g.in_edges(s,data='var')])
            == lpSum([t[2] for t in g.edges(s,data='var')]))

    m.solve() 

    h = nx.MultiDiGraph()
    addvals(df)

    for row in df[df.Val > 0.5].itertuples():
        h.add_edge(row.From, row.To, word=row.Word)

    # h.add_edge('end','start')
    # res = [h[f][t][k]['word'] for f, t, k in list(nx.eulerian_circuit(h, 'start', True))[1:-2]]

    h.add_edge('end', 'start')
    h = h.subgraph(list(nx.weakly_connected_components(h))[0])
    res = [h[f][t][k]['word'] for f, t, k in list(nx.eulerian_circuit(h, 'start', True))[1:-2]]

    return res

#9 結果をExcelに貼り付け
def Mk_Excel(res,Rep_Poke_Df,Poke_Df):

    Comp_List = []

    for r in res:
        subDf = Rep_Poke_Df[Rep_Poke_Df['PokeName'] == r]
        Num = subDf.iat[0,0]
        sub2Df = Poke_Df[Poke_Df['Zukan'] == Num]
        Poke = sub2Df.iat[0,1]
        t = (Poke[0],Poke[-1],Num,Poke)
        Comp_List.append(t)

    Comp_Df = pd.DataFrame(Comp_List,columns = ['Start','End','No','Name'])

    Comp_Df.to_excel(setup.ExcelFile)

    GraphRow = len(Comp_Df)
    GraphCol = len(Comp_Df.columns)

    book = openpyxl.load_workbook(setup.ExcelFile)
    sheet = book.worksheets[0]
    CellRangeCol = sheet.iter_rows(min_row=1, max_row= 1, min_col=1, max_col= GraphCol+1)
    CellRangeRow = sheet.iter_rows(min_row=2, max_row= GraphRow+1, min_col=1, max_col= 1)
    CellAllRange = sheet.iter_rows(min_row=2, max_row= GraphRow+1, min_col=2, max_col= GraphCol+1)

    ColColor = PatternFill(patternType = 'solid',fgColor = 'FFFF00') #黄色
    RowColor = PatternFill(patternType = 'solid',fgColor = 'FABF8F') #橙色
    GraphColor = PatternFill(patternType = 'solid',fgColor = 'DAEEF3') #薄青

    Header_Font = Font(name='Meiryo UI' , size=12 , color='000000')
    Content_Font = Font(name='Meiryo UI' , size=11 , color='000000')

    side = Side(style='thin', color='000000')

    for r in CellRangeCol:
        for cell in r:
            cell.fill = ColColor
            cell.font = Header_Font

    for r in CellRangeRow:
        for cell in r:
            cell.fill = RowColor
            cell.font = Header_Font
    
    for r in CellAllRange:
        for cell in r:
            cell.border = Border(left=side , right=side , top=side , bottom=side)
            cell.font = Content_Font
            if cell.row % 2 == 0:
                cell.fill = GraphColor
            
    book.save(setup.ExcelFile)

#0 処理開始MSG
now = datetime.datetime.now()
Log = open(setup.LogFile,'a')
Log.write('------------------------------------------------------\n')
Log.write('#0_処理実行 ' + str(now.strftime('%Y年%m月%d日 %H : %M : %S')) + '\n')
print('------------------------------------------------------')
print('処理実行 ' + str(now.strftime('%Y年%m月%d日 %H : %M : %S')))

#1 前回ファイルの削除
Log.write('#1_前回ファイル削除_実行\n')
print('#1_前回ファイル削除')
Rem_File()
Log.write('#1_前回ファイル削除_完了\n')

#2 ポケモン名の取得
Log.write('#2_ポケモン名取得_実行\n')
print('#2_ポケモン名取得')
Poke_Df = Get_PokeName()
Log.write('#2_ポケモン名取得_完了\n')

#3 最後に「ん」の付くポケモンを除外
Log.write('#3_最後に「ん」の付くポケモン除外_実行\n')
print('#3_最後に「ん」の付くポケモン除外')
# Rep_Poke_Df = Rem_Poke_LastN(Poke_Df)
Log.write('#3_最後に「ん」の付くポケモン除外_完了\n')

#4 最後の文字が特殊文字の場合にカタカナに変更
Log.write('#4_特殊文字をカタカナに変更_実行\n')
print('#4_特殊文字をカタカナに変更')
# Rep_Poke_Df = Rep_SpecialtoChara(Rep_Poke_Df)
Rep_Poke_Df = Rep_SpecialtoChara(Poke_Df)
Log.write('#4_特殊文字をカタカナに変更_完了\n')

#5 最後の文字がハイフンの場合の処理
Log.write('#5_ハイフン文字変更_実行\n')
print('#5_ハイフン文字変更')
Rep_Poke_Df = Rep_HyphentoChara(Rep_Poke_Df,setup.Hyphen_JD)
Log.write('#5_ハイフン文字変更_完了\n')

#6 最後の文字が小文字の場合に大文字に変更
Log.write('#6_小文字を大文字に変更_実行\n')
print('#6_小文字を大文字に変更')
Rep_Poke_Df = Rep_SmalltoBig(Rep_Poke_Df)
Log.write('#6_小文字を大文字に変更_完了\n')

#7 濁点を考慮しない場合の濁点除去処理
if setup.Dakuten_JD:
    Log.write('#7_濁点を考慮しない場合の濁点除去処理_実行\n')
    print('#7_濁点を考慮しない場合の濁点除去処理')
    Rep_Poke_Df = Rem_Dakuten(Rep_Poke_Df)
    Log.write('#7_濁点を考慮しない場合の濁点除去処理_完了\n')
else:
    print('#7_濁点を考慮しない場合の濁点除去処理_スキップ')
    Log.write('#7_濁点を考慮しない場合の濁点除去処理_スキップ\n')

#8 しりとり作成
Log.write('#8_しりとり作成_実行\n')
print('#8_しりとり作成')
res = Mk_Shiritori(Rep_Poke_Df,Poke_Df)
Log.write('#8_しりとり作成_完了\n')

#9 結果をExcelに貼り付け
Log.write('#9_Excel作成_実行\n')
print('#9_Excel作成')
Mk_Excel(res,Rep_Poke_Df,Poke_Df)
Log.write('#9_Excel作成_完了\n')

#10 終了処理
now = datetime.datetime.now()
print('処理完了 ' + str(now.strftime('%Y年%m月%d日 %H : %M : %S')))
print('------------------------------------------------------')
Log.write('#_処理完了 ' + str(now.strftime('%Y年%m月%d日 %H : %M : %S')) + '\n')
Log.write('------------------------------------------------------\n')


